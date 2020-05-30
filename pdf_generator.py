#encoding=utf8
import os
import sys
from pyPdf import PdfFileWriter, PdfFileReader
import StringIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from subprocess import call

'''
	No s'ha de fer contracte de voluntariat
	#fabio pinto, natalie, dir, yasmina garcia, sandra granada, hector muñoz delgado, nuria santoa, invira luz, gustavo kindelan, 
	arman, danislay.

'''

def get_info(num,file,letters):
	import pandas as pd
	from openpyxl import load_workbook
	
	wb = load_workbook(file)
	sheet = wb.get_sheet_by_name('Hoja1')
	
	info = []
	# D(nom), E(cognom), G(tel), I(dni)
	for let in letters:
		l = str(let)+str(num)
		if (sheet[str(l)].value == None or sheet[str(l)].value == 'none' or sheet[str(l)].value == 'None'):
			info.append('0')
		else:
			info.append(str(sheet[str(let)+str(num)].value))	
	return info
	
def import_dni():
	'''import pandas as pd
	from openpyxl import load_workbook
	file_in = 'input_per_omplir.xlsx'
	wb = load_workbook(file_in)
	sheet = wb.get_sheet_by_name('Hoja1')

	for i in range(6,193):
		print sheet[str('S')+str(i)].value

	return 0
	'''
	import xlwt
	book = xlwt.Workbook()
	sh = book.add_sheet('Hoja2')
   	
	cont = 0
	file_in = 'input_per_omplir.xlsx'
	letters_in = ['D','H']

	file_out = 'full_personal_per_omplir.xlsx'
	letters_out = ['P','R']

	string = ''
	# del full_input agafem tota la info personal i la comparem amb la d'entrada (mobils) i afegim DNI
	#with open("dni.txt","w") as output_file:
	for i in range(6,167):
		info_in = get_info(i,file_in,letters_in)
		for x in range(6,195):
			info_out = get_info(x,file_in,letters_out)
			
			if (str(info_in[1]) == '0' or str(info_out[0]) == '0'):
				print info_in[0],' 0'
				break
			else:
				if (str(info_out[0]) == str(info_in[1])):
					print str(info_in[0]),str(info_out[1]).upper()
					break

		#output_file.write(str(info_out[1])+'')


			
			
	
	


	# en el full fullpersonal_per_omplir a la columna 5 o F, escriure el dni
	
def treat_xlsx():
	import pandas as pd
	from openpyxl import load_workbook
	
	#os.chdir('/Users/poooool/Desktop')

	file = 'google.xlsx'

	wb = load_workbook(file)
	sheet = wb.get_sheet_by_name('Hoja1')
	
	letters = ['B','C','A']
	
	s1 = ''
	
	# de num - 1 a num + 1

	#for letter in letters:
	for i in range(4,190):
		#s1 += str(sheet[str(letter)+str(i)].value) + ' | '
		#s1 += str(sheet['D'+str(i)].value) + ' | '
		if (sheet['I'+str(i)].value == None):
			print 'NONNNE'
			pass
		else:
			print sheet['I'+str(i)].value

	s1 += '\n'

	print s1

def make_pdf_1(name='', surname1='',dni=''):

	packet = StringIO.StringIO()

	can = canvas.Canvas(packet, pagesize=letter)
	can.drawString(290, 570, str(name)+' '+str(surname1))
	can.drawString(210, 540, str(dni))
	can.drawString(230, 370, str(name)+' '+str(surname1))
	can.drawString(230, 137, str(name)+' '+str(surname1))

	can.save()

	#move to the beginning of the StringIO buffer
	packet.seek(0)
	new_pdf = PdfFileReader(packet)
	
	# read your existing PDF
	existing_pdf = PdfFileReader(file("contracte_voluntariat.pdf", "rb"))
	output = PdfFileWriter()
	
	# add the "watermark" (which is the new pdf) on the existing page
	#print 'hola '+str(existing_pdf.getNumPages())

	page = existing_pdf.getPage(0)

	page.mergePage(new_pdf.getPage(0))
	output.addPage(page)
	filename = str(dni)+'file'+str(0)+'.pdf'
	
	# finally, write "output" to a real file
	outputStream = file(filename, "wb")
	output.write(outputStream)
	outputStream.close()

	#make_pdf(name,surname1,surname2,dni,n_page + 1)

	return filename

def make_pdf_2(dni=''):
	packet = StringIO.StringIO()

	can = canvas.Canvas(packet, pagesize=letter)
	can.drawString(257, 568, '28 de Julio de 2018')
	
	can.save()

	#move to the beginning of the StringIO buffer
	packet.seek(0)
	new_pdf = PdfFileReader(packet)
	
	# read your existing PDF
	existing_pdf = PdfFileReader(file("contracte_voluntariat.pdf", "rb"))
	output = PdfFileWriter()
	
	# add the "watermark" (which is the new pdf) on the existing page
	#print 'hola '+str(existing_pdf.getNumPages())

	page = existing_pdf.getPage(1)

	page.mergePage(new_pdf.getPage(0))
	output.addPage(page)
	filename = str(dni)+'file'+str(1)+'.pdf'
	
	# finally, write "output" to a real file
	outputStream = file(filename, "wb")
	output.write(outputStream)
	outputStream.close()



	return filename	

def make_pdf_3(name='',surname1='',dni=''):
	packet = StringIO.StringIO()

	can = canvas.Canvas(packet, pagesize=letter)
	can.drawString(360, 122, str(name)+' '+str(surname1))
	can.drawString(365, 105, str(dni))
	can.drawString(375, 640, '28')
	can.drawString(480, 640, '28')

	can.save()

	#move to the beginning of the StringIO buffer
	packet.seek(0)
	new_pdf = PdfFileReader(packet)
	
	# read your existing PDF
	existing_pdf = PdfFileReader(file("contracte_voluntariat.pdf", "rb"))
	output = PdfFileWriter()
	
	# add the "watermark" (which is the new pdf) on the existing page
	#print 'hola '+str(existing_pdf.getNumPages())

	page = existing_pdf.getPage(5)

	page.mergePage(new_pdf.getPage(0))
	output.addPage(page)
	filename = str(dni)+'file'+str(5)+'.pdf'
	
	# finally, write "output" to a real file
	outputStream = file(filename, "wb")
	output.write(outputStream)
	outputStream.close()



	return filename

def make_all_pdf(name='', surname1='', dni=''):
	os.chdir('/Users/poooool/Desktop/RBF')
	make_pdf_1(name,surname1,dni)
	make_pdf_2(dni)
	make_pdf_3(name,surname1,dni)

	packet = StringIO.StringIO()

	can = canvas.Canvas(packet, pagesize=letter)
	can.drawString(0, 0, ' ')

	can.save()

	#move to the beginning of the StringIO buffer
	packet.seek(0)
	new_pdf = PdfFileReader(packet)
	
	# read your existing PDF
	os.chdir('/Users/poooool/Desktop/RBF/')
	existing_pdf = PdfFileReader(file("contracte_voluntariat.pdf", "rb"))
	output = PdfFileWriter()
	
	# add the "watermark" (which is the new pdf) on the existing page
	#print 'hola '+str(existing_pdf.getNumPages())
	for i in range(2,existing_pdf.getNumPages()-1):
		page = existing_pdf.getPage(i)

		page.mergePage(new_pdf.getPage(0))
		output.addPage(page)
		filename = str(dni)+'file'+str(i)+'.pdf'
		# finally, write "output" to a real file
		outputStream = file(filename, "wb")
		output.write(outputStream)
		outputStream.close()

	from PyPDF2 import PdfFileMerger

	pdfs = ['file0.pdf', 'file1.pdf', 'file2.pdf', 'file3.pdf','file4.pdf','file5.pdf']
	merger = PdfFileMerger()
	for pdf in pdfs:
		merger.append(str(dni)+str(pdf))
	
	#output_filename = "contractes-voluntariat-personal-"+str(name)+'-'+str(dni)+".pdf"
	output_filename = "contracte-voluntariat-animacio-"+str(name)+'-'+str(dni)+".pdf"

	os.chdir('/Users/poooool/Desktop/RBF/contractes-voluntariat-personal')
	#os.chdir('/Users/poooool/Desktop/RBF/animacio')

	merger.write(output_filename)
	os.chdir('/Users/poooool/Desktop/RBF/')

	for p in pdfs:
		call(['rm',str(dni)+str(p)])


if __name__ == '__main__':
	
	#treat_xlsx()
	
	reload(sys)
	sys.setdefaultencoding('utf8')
	#import_dni()
	letters = ['A','B','C']
	letters2 = ['H','I']

	for i in range(5,8):
		info = get_info(i,'personal.xlsx',letters)
		if str(info[2]) == '0':
			info[2] = ''
			print 'en ',str(info[0]),' no te dni'
		
		make_all_pdf(name=str(info[0]),surname1=str(info[1]),dni = str(info[2]))